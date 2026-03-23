# %% [markdown]
# # Data Profiling Notebook — Template
#
# **Purpose:** Profile raw source files and produce a portable JSON bundle
# for downstream Silver layer transformation design.
#
# **How to use:**
# 1. Upload source files to Lakehouse `Files/raw/`
# 2. Update CONFIG in Cell 2 with your file names and join keys
# 3. Run All
# 4. Download results from `Files/output/`
#
# **Checks performed:** Encoding, BOM, header shifts, duplicate headers,
# type inference, null conventions, PII detection, date formats, currency
# formatting, cross-file join key matching, Excel hidden sheets, and more.

# %% Cell 1 — Install dependencies
# Only needed if your Fabric Environment doesn't include openpyxl
%pip install openpyxl==3.1.5

# %% Cell 2 — Configuration
# =====================================================================
# EDIT THIS SECTION for your project
# =====================================================================
CONFIG = {
    # Source file names (must match exactly, case-sensitive)
    "file1_name": "{{FILE_1_NAME}}.csv",
    "file2_name": "{{FILE_2_NAME}}.csv",
    "file3_name": "{{FILE_3_NAME}}.xlsx",

    # Cross-file join keys
    "join_key_file1": "{{JOIN_KEY_COL_FILE1}}",
    "join_key_file2": "{{JOIN_KEY_COL_FILE2}}",
    "join_key_file3": "{{JOIN_KEY_COL_FILE3}}",

    # Known PII columns (for flagging)
    "pii_columns": ["Name", "SSN", "Email", "Phone", "Address", "Birth_Date", "Salary"],

    # Output location (Fabric Lakehouse path)
    "output_folder": "/lakehouse/default/Files/output",

    # Raw files location
    "raw_folder": "/lakehouse/default/Files/raw",

    # Encodings to try (in order)
    "encodings_to_try": ["utf-8-sig", "utf-8", "cp1252", "latin-1"],

    # Header shift: max rows to check for metadata before actual headers
    "header_shift_max_rows": 5,
}

# Verify files exist
import os
raw = CONFIG["raw_folder"]
for key in ["file1_name", "file2_name", "file3_name"]:
    path = os.path.join(raw, CONFIG[key])
    exists = os.path.exists(path)
    print(f"  {'[OK]' if exists else '[!!]'} {CONFIG[key]}: {'found' if exists else 'NOT FOUND'}")

# %% Cell 3 — Imports and helper functions
import pandas as pd
import numpy as np
import json
import re
import hashlib
from datetime import datetime

# For XLSX hidden sheet detection
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not available. XLSX hidden sheet detection disabled.")

RUN_TIMESTAMP = datetime.now().isoformat()
FINDINGS = []
ALL_RESULTS = []

def add_finding(severity, section, message, column=None, details=None):
    """Record a profiling finding."""
    finding = {
        "severity": severity,
        "section": section,
        "message": message,
    }
    if column:
        finding["column"] = column
    if details:
        finding["details"] = details
    FINDINGS.append(finding)
    symbol = {"CRITICAL": "[!!]", "WARNING": "[!]", "INFO": "[i]"}
    print(f"  {symbol.get(severity, '[?]')} {severity}: {message}")

def detect_encoding(filepath):
    """Try multiple encodings, return the first that works."""
    for enc in CONFIG["encodings_to_try"]:
        try:
            pd.read_csv(filepath, encoding=enc, nrows=2)
            return enc
        except (UnicodeDecodeError, Exception):
            continue
    return "utf-8"

def detect_header_shift(filepath, encoding):
    """Check if CSV starts with metadata rows before actual headers."""
    try:
        preview = pd.read_csv(filepath, encoding=encoding, header=None, nrows=CONFIG["header_shift_max_rows"])
        for i, row in preview.iterrows():
            non_null = row.dropna()
            if len(non_null) > len(row) * 0.5:
                if i > 0:
                    add_finding("CRITICAL", filepath, f"Header shift detected: actual headers at row {i+1}, first {i} rows are metadata", details=f"Use skiprows={i} when reading")
                return i
        return 0
    except Exception:
        return 0

def check_bom(filepath):
    """Check for UTF-8 BOM marker."""
    with open(filepath, "rb") as f:
        if f.read(3) == b'\xef\xbb\xbf':
            add_finding("INFO", filepath, "BOM detected -- will be stripped")
            return True
    return False

def check_duplicate_headers(columns):
    """Detect and report duplicate column names."""
    seen = {}
    duplicates = []
    for i, col in enumerate(columns):
        if col in seen:
            duplicates.append((col, seen[col], i))
        else:
            seen[col] = i
    return duplicates

def disambiguate_columns(columns):
    """Rename duplicate columns with positional suffixes."""
    counts = {}
    result = []
    for col in columns:
        if col in counts:
            counts[col] += 1
            result.append(f"{col}_{counts[col]}")
        else:
            counts[col] = 1
            result.append(col)
    # Fix first occurrence if there were duplicates
    final = []
    dup_set = {col for col, cnt in counts.items() if cnt > 1}
    first_seen = {}
    for col in result:
        base = col.rsplit("_", 1)[0] if "_" in col and col.rsplit("_", 1)[1].isdigit() else col
        if base in dup_set and base not in first_seen:
            first_seen[base] = True
            final.append(f"{base}_1")
        else:
            final.append(col)
    return final

def infer_type(series):
    """Infer the likely data type of a pandas Series."""
    non_null = series.dropna().astype(str).str.strip()
    non_null = non_null[non_null != ""]
    if len(non_null) == 0:
        return "empty"

    # Check boolean
    bool_vals = {"yes", "no", "y", "n", "true", "false", "1", "0"}
    if set(non_null.str.lower().unique()).issubset(bool_vals):
        return "boolean"

    # Check date patterns
    date_patterns = [
        r'^\d{1,2}/\d{1,2}/\d{2,4}$',
        r'^\d{4}-\d{2}-\d{2}',
        r'^\d{1,2}-\d{1,2}-\d{2,4}$',
    ]
    for pat in date_patterns:
        if non_null.str.match(pat).mean() > 0.8:
            return "date"

    # Check numeric (with currency symbols)
    cleaned = non_null.str.replace(r'[$,]', '', regex=True)
    try:
        pd.to_numeric(cleaned, errors='raise')
        if cleaned.str.contains(r'\.').any():
            return "decimal"
        return "integer"
    except (ValueError, TypeError):
        pass

    return "string"

def check_pii(col_name, series):
    """Check for PII patterns in column values."""
    non_null = series.dropna().astype(str)

    # SSN pattern
    if non_null.str.match(r'^\d{3}-\d{2}-\d{4}$').any():
        add_finding("CRITICAL", "PII", f"SSN pattern detected in column '{col_name}'", column=col_name)
        return True

    # Email pattern
    if non_null.str.contains(r'@.*\.', regex=True).mean() > 0.5:
        add_finding("WARNING", "PII", f"Email addresses detected in column '{col_name}'", column=col_name)
        return True

    # Check column name against known PII columns
    pii_keywords = [p.lower() for p in CONFIG["pii_columns"]]
    if any(kw in col_name.lower() for kw in pii_keywords):
        add_finding("WARNING", "PII", f"Column '{col_name}' name matches PII keyword list", column=col_name)
        return True

    return False

def profile_column(source_file, col_name, col_idx, series):
    """Profile a single column and return results dict."""
    total = len(series)
    null_count = series.isna().sum()
    blank_count = (series.astype(str).str.strip() == "").sum() - null_count
    na_count = series.astype(str).str.lower().isin(["n/a", "na", "none", "-", "null"]).sum()

    nullish_total = null_count + max(0, blank_count) + na_count
    nullish_pct = round(nullish_total / total * 100, 1) if total > 0 else 0

    non_null = series.dropna()
    distinct = non_null.nunique()
    inferred = infer_type(series)

    issues = []

    # Whitespace padding
    if non_null.astype(str).str.contains(r'^\s|\s$', regex=True).any():
        issues.append("whitespace_padding")
        add_finding("WARNING", source_file, f"Whitespace padding in '{col_name}'", column=col_name)

    # Currency symbols in numeric columns
    if non_null.astype(str).str.contains(r'[\$,]', regex=True).any() and inferred in ("decimal", "integer"):
        issues.append("currency_symbols")

    # Leading zeros
    if inferred == "integer":
        str_vals = non_null.astype(str)
        if str_vals.str.match(r'^0\d+').any():
            issues.append("leading_zeros")
            add_finding("WARNING", source_file, f"Leading zeros in '{col_name}' — keep as STRING to preserve", column=col_name)

    # PII check
    if check_pii(col_name, series):
        issues.append("pii_detected")

    result = {
        "source_file": source_file,
        "column_name": col_name,
        "column_index": col_idx,
        "inferred_type": inferred,
        "nullish_pct": nullish_pct,
        "distinct_count": int(distinct),
        "issues": issues,
    }

    ALL_RESULTS.append(result)
    return result

print("Helpers loaded.")

# %% Cell 4 — Load File 1 (CSV)
f1_path = os.path.join(CONFIG["raw_folder"], CONFIG["file1_name"])
check_bom(f1_path)
f1_encoding = detect_encoding(f1_path)
f1_skip = detect_header_shift(f1_path, f1_encoding)

if f1_encoding != "utf-8-sig" and f1_encoding != "utf-8":
    add_finding("WARNING", CONFIG["file1_name"], f"Non-UTF-8 encoding detected: {f1_encoding}")

pdf1 = pd.read_csv(f1_path, encoding=f1_encoding, skiprows=f1_skip, dtype=str)
f1_original_headers = list(pdf1.columns)

# Check for duplicate headers
f1_dupes = check_duplicate_headers(pdf1.columns)
if f1_dupes:
    add_finding("CRITICAL", CONFIG["file1_name"],
        f"Duplicate column headers found: {len(f1_dupes)} duplicates",
        details=str([(d[0], f"positions {d[1]} and {d[2]}") for d in f1_dupes]))
    pdf1.columns = disambiguate_columns(pdf1.columns)

print(f"File 1: {CONFIG['file1_name']} — {len(pdf1)} rows, {len(pdf1.columns)} columns, encoding={f1_encoding}")

# %% Cell 5 — Load File 2 (CSV)
f2_path = os.path.join(CONFIG["raw_folder"], CONFIG["file2_name"])
check_bom(f2_path)
f2_encoding = detect_encoding(f2_path)
f2_skip = detect_header_shift(f2_path, f2_encoding)

if f2_encoding != "utf-8-sig" and f2_encoding != "utf-8":
    add_finding("WARNING", CONFIG["file2_name"], f"Non-UTF-8 encoding detected: {f2_encoding}")

pdf2 = pd.read_csv(f2_path, encoding=f2_encoding, skiprows=f2_skip, dtype=str)
f2_original_headers = list(pdf2.columns)

f2_dupes = check_duplicate_headers(pdf2.columns)
if f2_dupes:
    add_finding("CRITICAL", CONFIG["file2_name"],
        f"Duplicate column headers found: {len(f2_dupes)} duplicates",
        details=str([(d[0], f"positions {d[1]} and {d[2]}") for d in f2_dupes]))
    pdf2.columns = disambiguate_columns(pdf2.columns)

print(f"File 2: {CONFIG['file2_name']} — {len(pdf2)} rows, {len(pdf2.columns)} columns, encoding={f2_encoding}")

# %% Cell 6 — Load File 3 (XLSX)
f3_path = os.path.join(CONFIG["raw_folder"], CONFIG["file3_name"])

# Hidden sheet detection
if HAS_OPENPYXL:
    wb = openpyxl.load_workbook(f3_path, read_only=True)
    visible_sheets = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]
    hidden_sheets = [s for s in wb.sheetnames if wb[s].sheet_state != "visible"]
    if hidden_sheets:
        add_finding("WARNING", CONFIG["file3_name"],
            f"Hidden sheets detected: {hidden_sheets}",
            details="Inspect hidden sheets for additional data")
    wb.close()
    print(f"  Sheets: {wb.sheetnames} (hidden: {hidden_sheets})")

pdf3 = pd.read_excel(f3_path, dtype=str)

# Check for formula wrappers (e.g., ="00123")
for col in pdf3.columns:
    if pdf3[col].astype(str).str.match(r'^=".*"$').any():
        add_finding("WARNING", CONFIG["file3_name"],
            f"Excel formula wrappers detected in '{col}' (e.g., =\"00123\")",
            column=col, details="Strip formula wrappers: remove leading = and surrounding quotes")

print(f"File 3: {CONFIG['file3_name']} — {len(pdf3)} rows, {len(pdf3.columns)} columns")

# %% Cell 7 — File Metadata Summary
print("=" * 70)
print("FILE METADATA SUMMARY")
print("=" * 70)
summary_data = [
    {"File": CONFIG["file1_name"], "Rows": len(pdf1), "Columns": len(pdf1.columns), "Encoding": f1_encoding, "Format": "CSV"},
    {"File": CONFIG["file2_name"], "Rows": len(pdf2), "Columns": len(pdf2.columns), "Encoding": f2_encoding, "Format": "CSV"},
    {"File": CONFIG["file3_name"], "Rows": len(pdf3), "Columns": len(pdf3.columns), "Encoding": "N/A (XLSX)", "Format": "XLSX"},
]
display(pd.DataFrame(summary_data))

# %% Cell 8 — Profile File 1
print("=" * 70)
print(f"PROFILING: {CONFIG['file1_name']}")
print("=" * 70)
for i, col in enumerate(pdf1.columns):
    profile_column(CONFIG["file1_name"], col, i, pdf1[col])
print(f"\n  Profiled {len(pdf1.columns)} columns from File 1")

# %% Cell 9 — Profile File 2
print("=" * 70)
print(f"PROFILING: {CONFIG['file2_name']}")
print("=" * 70)
for i, col in enumerate(pdf2.columns):
    profile_column(CONFIG["file2_name"], col, i, pdf2[col])
print(f"\n  Profiled {len(pdf2.columns)} columns from File 2")

# %% Cell 10 — Profile File 3
print("=" * 70)
print(f"PROFILING: {CONFIG['file3_name']}")
print("=" * 70)
for i, col in enumerate(pdf3.columns):
    profile_column(CONFIG["file3_name"], col, i, pdf3[col])
print(f"\n  Profiled {len(pdf3.columns)} columns from File 3")

# %% Cell 11 — Cross-File Validation
print("=" * 70)
print("CROSS-FILE VALIDATION")
print("=" * 70)

# Extract join keys
key1 = CONFIG["join_key_file1"]
key2 = CONFIG["join_key_file2"]
key3 = CONFIG["join_key_file3"]

set1 = set(pdf1[key1].dropna().astype(str).str.strip()) if key1 in pdf1.columns else set()
set2 = set(pdf2[key2].dropna().astype(str).str.strip()) if key2 in pdf2.columns else set()
set3 = set(pdf3[key3].dropna().astype(str).str.strip()) if key3 in pdf3.columns else set()

in_all = set1 & set2 & set3
only_1 = set1 - set2 - set3
only_2 = set2 - set1 - set3
only_3 = set3 - set1 - set2

print(f"  File 1 unique keys: {len(set1)}")
print(f"  File 2 unique keys: {len(set2)}")
print(f"  File 3 unique keys: {len(set3)}")
print(f"  In all 3 files: {len(in_all)}")
print(f"  Only in File 1: {len(only_1)}")
print(f"  Only in File 2: {len(only_2)}")
print(f"  Only in File 3: {len(only_3)}")

if only_3:
    add_finding("WARNING", "Cross-File",
        f"{len(only_3)} keys in File 3 not found in Files 1 or 2 (possible manual entries)",
        details=f"Sample: {list(only_3)[:10]}")

# Leading zero inconsistency
has_leading_zero_1 = any(k.startswith("0") for k in set1) if set1 else False
has_leading_zero_3 = any(k.startswith("0") for k in set3) if set3 else False
if has_leading_zero_1 != has_leading_zero_3:
    add_finding("CRITICAL", "Cross-File",
        "Leading zero inconsistency between files — join keys won't match without normalization",
        details=f"File 1 has leading zeros: {has_leading_zero_1}, File 3: {has_leading_zero_3}")

# Duplicate keys (rehire detection)
for name, pdf, key in [(CONFIG["file1_name"], pdf1, key1), (CONFIG["file2_name"], pdf2, key2)]:
    if key in pdf.columns:
        dupes = pdf[key].dropna().astype(str).str.strip()
        dup_keys = dupes[dupes.duplicated()].unique()
        if len(dup_keys) > 0:
            add_finding("WARNING", name,
                f"Duplicate join keys found: {len(dup_keys)} keys appear more than once (possible rehires or multi-row records)",
                details=f"Sample: {list(dup_keys[:5])}")

# %% Cell 12 — Null Convention Analysis
print("=" * 70)
print("NULL CONVENTION ANALYSIS")
print("=" * 70)

null_map = []
for r in ALL_RESULTS:
    col_data = None
    if r["source_file"] == CONFIG["file1_name"]:
        col_data = pdf1[r["column_name"]] if r["column_name"] in pdf1.columns else None
    elif r["source_file"] == CONFIG["file2_name"]:
        col_data = pdf2[r["column_name"]] if r["column_name"] in pdf2.columns else None
    elif r["source_file"] == CONFIG["file3_name"]:
        col_data = pdf3[r["column_name"]] if r["column_name"] in pdf3.columns else None

    if col_data is None:
        continue

    str_vals = col_data.astype(str).str.strip()
    conventions = {
        "null": int(col_data.isna().sum()),
        "empty_string": int((str_vals == "").sum() - col_data.isna().sum()),
        "na_text": int(str_vals.str.lower().isin(["n/a", "na", "none"]).sum()),
        "dash": int((str_vals == "-").sum()),
        "zero": int((str_vals == "0").sum()),
    }

    dominant = max(conventions, key=conventions.get) if any(v > 0 for v in conventions.values()) else "none"

    null_map.append({
        "source_file": r["source_file"],
        "column_name": r["column_name"],
        "null_count": conventions["null"],
        "empty_count": conventions["empty_string"],
        "na_text_count": conventions["na_text"],
        "dash_count": conventions["dash"],
        "zero_count": conventions["zero"],
        "dominant_convention": dominant,
        "silver_recommendation": f"Map {dominant} to NULL" if dominant != "none" else "No nulls detected",
    })

# Display as HTML table with color coding
null_df = pd.DataFrame(null_map)
display(null_df)

# %% Cell 13 — Summary Report (HTML)
print("=" * 70)
print("GENERATING HTML REPORT")
print("=" * 70)

critical_count = sum(1 for f in FINDINGS if f["severity"] == "CRITICAL")
warning_count = sum(1 for f in FINDINGS if f["severity"] == "WARNING")
info_count = sum(1 for f in FINDINGS if f["severity"] == "INFO")

html = f"""
<html><head><style>
body {{ font-family: -apple-system, sans-serif; max-width: 1200px; margin: 0 auto; padding: 20px; }}
.kpi {{ display: inline-block; padding: 20px 30px; margin: 10px; border-radius: 8px; text-align: center; color: white; }}
.kpi-critical {{ background: #DC2626; }}
.kpi-warning {{ background: #F59E0B; }}
.kpi-info {{ background: #3B82F6; }}
.kpi-total {{ background: #1E293B; }}
.kpi h2 {{ margin: 0; font-size: 36px; }}
.kpi p {{ margin: 5px 0 0; font-size: 14px; }}
table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; font-size: 13px; }}
th {{ background: #1E293B; color: white; }}
tr:nth-child(even) {{ background: #f9f9f9; }}
.severity-CRITICAL {{ color: #DC2626; font-weight: bold; }}
.severity-WARNING {{ color: #D97706; font-weight: bold; }}
.severity-INFO {{ color: #3B82F6; }}
</style></head><body>
<h1>Data Profiling Report</h1>
<p>Generated: {RUN_TIMESTAMP}</p>
<div>
  <div class="kpi kpi-total"><h2>{len(ALL_RESULTS)}</h2><p>Columns Profiled</p></div>
  <div class="kpi kpi-critical"><h2>{critical_count}</h2><p>Critical</p></div>
  <div class="kpi kpi-warning"><h2>{warning_count}</h2><p>Warnings</p></div>
  <div class="kpi kpi-info"><h2>{info_count}</h2><p>Info</p></div>
</div>
<h2>Findings</h2>
<table><tr><th>Severity</th><th>Section</th><th>Column</th><th>Message</th><th>Details</th></tr>
"""
for f in sorted(FINDINGS, key=lambda x: {"CRITICAL": 0, "WARNING": 1, "INFO": 2}.get(x["severity"], 3)):
    html += f'<tr><td class="severity-{f["severity"]}">{f["severity"]}</td>'
    html += f'<td>{f["section"]}</td>'
    html += f'<td>{f.get("column", "")}</td>'
    html += f'<td>{f["message"]}</td>'
    html += f'<td>{f.get("details", "")}</td></tr>'
html += "</table></body></html>"

# Save HTML report
os.makedirs(CONFIG["output_folder"], exist_ok=True)
html_path = os.path.join(CONFIG["output_folder"], "profiling_report.html")
with open(html_path, "w", encoding="utf-8") as f:
    f.write(html)
print(f"  Saved: profiling_report.html")

# Display inline
from IPython.display import HTML, display as ipy_display
ipy_display(HTML(html))

# %% Cell 14 — Save Results to Delta Tables
print("=" * 70)
print("SAVING TO DELTA TABLES")
print("=" * 70)

results_pdf = pd.DataFrame(ALL_RESULTS)
for col in results_pdf.columns:
    results_pdf[col] = results_pdf[col].apply(
        lambda x: json.dumps(x) if isinstance(x, (dict, list)) else x
    )

results_sdf = spark.createDataFrame(results_pdf.astype(str))
results_sdf.write.mode("overwrite").saveAsTable("data_profiling_results")
print(f"  Saved {len(ALL_RESULTS)} column profiles to Tables/data_profiling_results")

if FINDINGS:
    findings_pdf = pd.DataFrame(FINDINGS).astype(str)
    findings_sdf = spark.createDataFrame(findings_pdf)
    findings_sdf.write.mode("overwrite").saveAsTable("data_profiling_findings")
    print(f"  Saved {len(FINDINGS)} findings to Tables/data_profiling_findings")

if null_map:
    null_pdf = pd.DataFrame(null_map)
    for col in null_pdf.columns:
        null_pdf[col] = null_pdf[col].apply(
            lambda x: json.dumps(x) if isinstance(x, (dict, list)) else x
        )
    null_sdf = spark.createDataFrame(null_pdf.astype(str))
    null_sdf.write.mode("overwrite").saveAsTable("data_profiling_null_map")
    print(f"  Saved {len(null_map)} null convention entries to Tables/data_profiling_null_map")

# %% Cell 15 — Export Portable Files (CSV + JSON)
print("=" * 70)
print("EXPORTING PORTABLE FILES")
print("=" * 70)

export_dir = CONFIG["output_folder"]

# CSV exports
results_export = pd.DataFrame(ALL_RESULTS)
for col in results_export.columns:
    results_export[col] = results_export[col].apply(
        lambda x: json.dumps(x) if isinstance(x, (dict, list)) else x
    )
results_export.to_csv(f"{export_dir}/profiling_results.csv", index=False)
print(f"  Saved: profiling_results.csv ({len(results_export)} rows)")

pd.DataFrame(FINDINGS).to_csv(f"{export_dir}/profiling_findings.csv", index=False)
print(f"  Saved: profiling_findings.csv ({len(FINDINGS)} findings)")

null_export = pd.DataFrame(null_map)
for col in null_export.columns:
    null_export[col] = null_export[col].apply(
        lambda x: json.dumps(x) if isinstance(x, (dict, list)) else x
    )
null_export.to_csv(f"{export_dir}/null_convention_map.csv", index=False)
print(f"  Saved: null_convention_map.csv ({len(null_export)} columns)")

# JSON bundle (comprehensive — share with Claude Code)
bundle = {
    "metadata": {
        "run_timestamp": RUN_TIMESTAMP,
        "file1": {"name": CONFIG["file1_name"], "rows": len(pdf1), "cols": len(pdf1.columns), "encoding": f1_encoding},
        "file2": {"name": CONFIG["file2_name"], "rows": len(pdf2), "cols": len(pdf2.columns), "encoding": f2_encoding},
        "file3": {"name": CONFIG["file3_name"], "rows": len(pdf3), "cols": len(pdf3.columns)},
    },
    "summary": {
        "total_columns_profiled": len(ALL_RESULTS),
        "critical_count": critical_count,
        "warning_count": warning_count,
        "info_count": info_count,
        "keys_in_all_files": len(in_all),
        "file1_only_keys": len(only_1),
        "file2_only_keys": len(only_2),
        "file3_only_keys": len(only_3),
    },
    "findings": FINDINGS,
    "null_convention_map": null_map,
    "column_profiles": ALL_RESULTS,
    "cross_file": {
        "file1_keys": len(set1),
        "file2_keys": len(set2),
        "file3_keys": len(set3),
        "overlap_all": len(in_all),
        "file3_manual_entries": list(set3 - set1 - set2) if len(set3 - set1 - set2) <= 20 else f"{len(set3 - set1 - set2)} codes",
    },
}

def convert_numpy(obj):
    if hasattr(obj, "item"):
        return obj.item()
    return str(obj)

json_path = f"{export_dir}/profiling_bundle.json"
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(bundle, f, indent=2, default=convert_numpy)
print(f"  Saved: profiling_bundle.json")

# File headers
headers_bundle = {
    "file1_headers": {"original": f1_original_headers, "renamed": list(pdf1.columns)},
    "file2_headers": {"original": f2_original_headers, "renamed": list(pdf2.columns)},
    "file3_headers": {"original": list(pdf3.columns), "renamed": list(pdf3.columns)},
}
with open(f"{export_dir}/file_headers.json", "w", encoding="utf-8") as f:
    json.dump(headers_bundle, f, indent=2)
print(f"  Saved: file_headers.json")

print(f"\n{'=' * 70}")
print("ALL EXPORTS COMPLETE")
print(f"{'=' * 70}")
print(f"\n  Download from: Lakehouse explorer > Files > output/")
print(f"  profiling_bundle.json  <- SHARE WITH CLAUDE CODE for Silver layer design")

# %% [markdown]
# ## Next Steps
#
# 1. **Download all files** from `Files/output/` in the Lakehouse explorer
# 2. **Review the HTML report** in your browser
# 3. **Share `profiling_bundle.json` with Claude Code** for automated Silver layer design
# 4. **Use `null_convention_map.csv`** to configure null normalization
# 5. **Share the HTML report** with stakeholders
